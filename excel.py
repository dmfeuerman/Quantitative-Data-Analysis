import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook


class EXCEL_WALKER:
    def __init__(self):
        self.start_path = None
        self.out_path = None
        
    def set_path(self, ticker):
        self.start_path = f"{ticker}_COMPLETE_DATA"
        self.out_path = f"stock_data/{ticker}/combined_data.xlsx"
        
        
        # Create the output directory if it doesn't exist
        os.makedirs(os.path.dirname(self.out_path), exist_ok=True)
        
    def walk(self):
        if self.start_path is None:
            print("Error: Path not set. Call set_path(ticker) first.")
            return
            
        # Set the starting directory
        directory = Path(self.start_path) 
        
        if not directory.exists():
            print(f"Error: Directory {self.start_path} does not exist.")
            return

        # Use rglob to recursively find all files matching the pattern "*.csv"
        csv_files = list(directory.rglob("*.csv"))
        
        if not csv_files:
            print(f"No CSV files found in {self.start_path}")
            return
            
        self._add_csv_to_excel_sheet(csv_files)
        

    def _add_csv_to_excel_sheet(self, csv_files):
        """
        Adds data from CSV files to an Excel file as new sheets.
        Creates the Excel file if it doesn't exist.
        """
        book = Workbook()
        # Remove the default blank sheet
        if 'Sheet' in book.sheetnames:
            del book['Sheet']
        print(f"Created new Excel workbook")

        # Process each CSV file
        for csv_file in csv_files:
            print(f"Processing CSV file: {csv_file}")
            new_sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
            
            # Read the CSV file into a pandas DataFrame
            df_new_sheet = pd.read_csv(csv_file)
            df_new_sheet = df_new_sheet.rename(columns= {
            'val': 'Value',
            'end': 'Period_End_Date',
            'start': 'Period_Start_Date',
            'accn': 'Accession_Number',
            'fy': 'Fiscal_Year',
            'fp': 'Fiscal_Period',
            'form': 'Form_Type',
            'filed': 'Filing_Date',
            'frame': 'Reporting_Frame'
        })
            
            # Check if the sheet name already exists and remove it to overwrite if needed
            if new_sheet_name in book.sheetnames:
                del book[new_sheet_name]

            # Create a new sheet and write the data
            ws = book.create_sheet(title=new_sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(df_new_sheet.columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row_data in enumerate(df_new_sheet.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            print(f"Added sheet '{new_sheet_name}'")
        
        # Save the workbook
        book.save(self.out_path)
        print(f"Successfully saved all sheets to {self.out_path}")
    

if __name__ == "__main__":
    excel = EXCEL_WALKER()
    excel.set_path("AAPL")  # Example ticker - change as needed
    excel.walk()